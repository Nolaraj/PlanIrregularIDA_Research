from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, RegularTextRun
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

import tkinter as tk
from tkinter import filedialog
import os
import sqlite3
import pandas as pd


def readInformationFile(file_path, sheet_name="Buildings"):
    # Read Excel without assuming headers
    raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Find the header row containing 'Parameters'
    try:
        header_idx = raw[raw.apply(lambda r: r.astype(str).str.contains('Parameters', case=False)).any(axis=1)].index[0]
    except:
        return
    # Read again using that row as header
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)

    # Clean column names
    df.columns = df.columns.astype(str).str.strip().str.replace(' ', '-', regex=False)

    # Set index to 'Parameters' (if exists), else use first column
    if 'Parameters' in df.columns:
        df.set_index('Parameters', inplace=True)
    else:
        df.set_index(df.columns[0], inplace=True)

    # Convert to dictionary
    return df.to_dict(orient='index')

def open_folder_dialog():
    root = tk.Tk()
    root.withdraw()

    folder_path = filedialog.askdirectory(
        title="Select the root folder containing the results"
    )

    return folder_path
def list_subfolders(root_folder):
    subfolders = []

    for dirpath, dirnames, filenames in os.walk(root_folder):
        for dirname in dirnames:
            subfolder_path = os.path.join(dirpath, dirname)
            subfolders.append(subfolder_path)

    return subfolders
def filePaths(Subfolders, FileName):
    ResultFiles = []
    for subfolder in subfolders:
        for files in os.listdir(subfolder):
            if files.endswith(FileName):
                Filepath = os.path.join(subfolder, files)
                ResultFiles.append(Filepath)

    return ResultFiles
def create_or_update_db(db_path, table_name, titles, values):
    # Connect to or create the database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Clean table name: replace spaces and invalid characters
    safe_table_name = "".join(ch if ch.isalnum() or ch == "_" else "-" for ch in table_name)

    # Clean and make unique column names
    clean_titles = []
    for i, t in enumerate(titles):
        clean = "".join(ch if ch.isalnum() or ch == "_" else "-" for ch in t)
        if not clean:
            clean = f"Column_{i}"
        # Handle duplicates
        while clean in clean_titles:
            clean += "_dup"
        clean_titles.append(clean)

    # Check if table exists
    cursor.execute(f"SELECT name FROM sqlite_master WHERE type='table' AND name=?", (safe_table_name,))
    table_exists = cursor.fetchone() is not None

    if not table_exists:
        print(f"🆕 Creating new table '{safe_table_name}' ...")
        columns_sql = ", ".join([f'"{col}" TEXT' for col in clean_titles])
        cursor.execute(f"""
            CREATE TABLE {safe_table_name} (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                {columns_sql}
            )
        """)
    else:
        print(f"✅ Table '{safe_table_name}' already exists, appending data...")

    # Insert rows
    placeholders = ", ".join(["?"] * len(clean_titles))
    columns_quoted = ", ".join([f'"{col}"' for col in clean_titles])
    cursor.execute(f"INSERT INTO {safe_table_name} ({columns_quoted}) VALUES ({placeholders})", values)

    conn.commit()
    conn.close()

def sqLite_writer(ResultFiles, OutputPath):
    Buildings_fromData = []
    Earthquakes_fromData = []
    ScaleFactors_fromData = []
    BaseCondtions_fromData = []



    for file in ResultFiles:
        SubjectCode = file.split(os.path.sep)[-5:-1]
        Earthquake = SubjectCode[0]
        ScaleFactor = SubjectCode[1]
        Building = SubjectCode[2]
        BaseCondition = SubjectCode[3]

        Earthquakes_fromData.append(Earthquake)
        ScaleFactors_fromData.append(ScaleFactor)
        Buildings_fromData.append(Building)
        BaseCondtions_fromData.append(BaseCondition)

        #Extracting to be written to the database file
        flat_titles = []
        flat_values = []
        flat_titles.extend(["Earthquake", "ScaleFactor", "Building", "BaseCondition"])
        flat_values.extend([Earthquake, ScaleFactor, Building, BaseCondition])

        def inputparaFlattening(parentDict, feature):
            reqDict = parentDict[feature]
            reqtitles = list(reqDict.keys())
            reqvalues = list(reqDict.values())
            flat_titles.extend(reqtitles)
            flat_values.extend(reqvalues)

        inputparaFlattening(Buildings_Info, Building)
        inputparaFlattening(Soils_Info, BaseCondition)
        inputparaFlattening(Earthquakes_Info, Earthquake)

        fObj = open(file, "r")
        lines = fObj.read()

        # Split the content into rows and cells based on tabs and newlines
        rows = lines.split('\n')
        if len(rows) == 1 and rows[0] == "":
            continue
        cells_in_rows = [row.split('\t') for row in rows]

        # Convert the cells to float values
        float_cells = []
        for row in cells_in_rows:
            float_row = []
            for cell_value in row:
                try:
                    float_value = float(cell_value)
                except ValueError:
                    float_value = cell_value
                float_row.append(float_value)
            float_cells.append(float_row)







        #Get nos of data rows excluding title
        titleRowsNo = 1
        for row_index, row in enumerate(float_cells, start=1):
            if row[0] == "":
                break
            RowsNos = row_index
        try:
            RowsNos = RowsNos - titleRowsNo
        except:
            RowsNos = 0

        #Creating for the titles data
        if RowsNos > 0:         #processing only if the file isnot empty and contains some of the parameteres
            table = float_cells
            # Extract headers
            headers = table[0]

            # Create dict to hold row indices for each column
            headers = [h.replace(" ", "-") for h in headers]
            column_titles = {header: [] for header in headers}
            column_values = {header: [] for header in headers}


            # Loop through data rows
            for row_index, row in enumerate(table[1:], start=1):  # start=1 to match row numbers as in Excel-style
                for col_index, value in enumerate(row):
                    if value != "" and value is not None:
                        column_name = headers[col_index]
                        column_titles[column_name].append(f"{column_name}_Level-{row_index}")
                        column_values[column_name].append(value)

            #Checking for the lengths of columns in both titles and data
            for col in headers:
                len_titles = len(column_titles[col])
                len_values = len(column_values[col])
                if len_titles != len_values:
                    print(f"⚠️ Mismatch mismatch in file {file}")

            # Flatten all titles and values in the same order
            for col in headers:
                titles = column_titles[col]
                values = column_values[col]

                # Ensure same length before extending
                flat_titles.extend(titles)
                flat_values.extend(values)

            # _____________________________________________________________Data Writing
            table_name = "IDA_Data"
            create_or_update_db(OutputPath, table_name, flat_titles, flat_values)


    def unique_preserve_order(seq):
        seen = set()
        return [x for x in seq if not (x in seen or seen.add(x))]

    Earthquakes_fromData = unique_preserve_order(Earthquakes_fromData)
    ScaleFactors_fromData = unique_preserve_order(ScaleFactors_fromData)
    Buildings_fromData = unique_preserve_order(Buildings_fromData)
    BaseCondtions_fromData = unique_preserve_order(BaseCondtions_fromData)

    return Earthquakes_fromData, ScaleFactors_fromData, Buildings_fromData, BaseCondtions_fromData


folderpath = open_folder_dialog()
subfolders = list_subfolders(folderpath)
FileName = "Result.txt"
ResultFiles = filePaths(subfolders, FileName)

Sheetname = "ResultAssembly"
rootpath = r"E:\Machine Learning Research\DataModelling"
OutputName =  "ResultAssembly"
OutputPath = os.path.join(rootpath, f'IDA_Data.db')

informationData = r"E:\Machine Learning Research\DataModelling\Modelling Information.xlsx"
Buildings_Info = readInformationFile(informationData, sheet_name="Buildings")
Earthquakes_Info = readInformationFile(informationData, sheet_name="Earthquakes")
Soils_Info = readInformationFile(informationData, sheet_name="Soils")


Earthquakes, ScaleFactors_fromData, Buildings, Soils = sqLite_writer(ResultFiles, OutputPath)
# for earthquake in Earthquakes:
#     print(earthquake)
# print(Earthquakes, ScaleFactors_fromData, Buildings, Soils)

from REQPY_Module import (REQPY_single, load_PEERNGA_record, plot_single_results,
                     save_results_as_at2, save_results_as_2col, save_results_as_1col)

from REQPY_Module import (REQPYrotdnn, load_PEERNGA_record, plot_rotdnn_results,
                          save_results_as_at2, save_results_as_2col, save_results_as_1col)
import REQPY_Module
import numpy as np
import matplotlib.pyplot as plt
import logging
import os
from collections import defaultdict
from typing import Tuple, List, Optional, Dict, Any
log = logging.getLogger(__name__)

import os
import shutil
import pandas as pd
from obspy import read
import numpy as np
import matplotlib.pyplot as plt
from RS_function import RS_function
import pandas as pd
import numpy as np
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference
import pandas as pd
from openpyxl.chart import ScatterChart
from openpyxl.styles import Font, Fill, Color, PatternFill  # Connect styles for text
from openpyxl.utils import get_column_letter, column_index_from_string
import openpyxl
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, RichTextProperties, Font, \
    RegularTextRun

import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.chart import ScatterChart, Reference, series
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
from openpyxl.chart import LineChart, Reference
import pandas as pd
from openpyxl.drawing.text import CharacterProperties, Font as DrawFont
from openpyxl.drawing.text import CharacterProperties, Font as DrawFont, ParagraphProperties
from openpyxl.chart.axis import ChartLines

def extract_eqname(filepath, EqNameOnly = True):
    # EqNameOnly provides only name of earthquake else whole earthquake info is passed out
    """Extract earthquake name info from the 2nd line of .AT2 file."""

    with open(filepath, 'r') as fp:
        next(fp)  # skip first line
        line2 = next(fp).strip().split(',')
        if len(line2) < 4:
            raise ValueError(f"Line 2 format incorrect in {filepath}. Expected Name, Date, Station, Component.")
        date_parts = line2[1].strip().split('/')
        if len(date_parts) < 3:
            raise ValueError(f"Date format incorrect in {filepath}. Expected MM/DD/YYYY.")
        year = date_parts[2]
        # eqname = f"{year}_{line2[0].strip()}_{line2[2].strip()}_comp_{line2[3].strip()}"
        if EqNameOnly:
            eqname = f"{year}_{line2[0].strip()}_{line2[2].strip()}"  # _comp_{line2[3].strip()}"
        else:
            eqname = f"{year}_{line2[0].strip()}_{line2[2].strip()}_comp_{line2[3].strip()}"

        return eqname


def EarthquakeGrouping(folderpath, display = True):
    """
    It groups all the earthquakes under the provided directory under the earthquake names
    :param folderpath:
    :return:
    """
    # Dictionary to group earthquake files
    # Format: { "base_name": [ (component_name, full_path), ... ] }
    earthquake_groups = defaultdict(list)

    for file in os.listdir(folderpath):
        filepath = os.path.join(folderpath, file)
        try:
            name = extract_eqname(filepath, EqNameOnly = False)

            # Extract base earthquake name (before "_comp_")
            if "_comp_" in name:
                base_name = name.split("_comp_")[0]
            else:
                base_name = name

            # Store tuple (component name, full path)
            earthquake_groups[base_name].append((name, filepath))

        except Exception as e:
            print(f"File error: {filepath}")
    # Display grouped results
    if display:
        for eq_name, files in earthquake_groups.items():
            print(f"\n🌍 Earthquake: {eq_name}")
            for comp_name, path in files:
                print(f"   └─ {comp_name}\n      📁 {path}")

    return earthquake_groups

def GrouptoFolder_FromSelectedinExcel(source_folders, excel_path, sheet_name, destination_root):
    """
    It basically extracts the row with file name under the
    Horizontal-1 Acc. Filename	 Horizontal-2 Acc. Filename	 Vertical Acc. Filename << under excel_path and sheet_name
    in excel sheet.
    It treats one row as one earthquake data
    Then those files are searched to source_folders directory.
    After that they are grouped in the folder named by function below.
    Then those earthquakes are then dispatched to the destination_root
    """

    # === READ EXCEL ===
    df = pd.read_excel(excel_path, sheet_name=sheet_name)



    def find_file_in_sources(filename):
        """Search for a file across multiple source folders."""
        for folder in source_folders:
            fullpath = os.path.join(folder, filename)
            if os.path.exists(fullpath):
                return fullpath
        return None

    # === PROCESS EACH ROW ===
    for idx, row in df.iterrows():
        # Collect .AT2 filenames in this row
        files_in_row = [str(v).strip() for v in row.values if isinstance(v, str) and v.strip().lower().endswith(".at2")]

        if not files_in_row:
            continue

        # --- Determine folder name from first available file ---
        eq_folder_name = None
        for target_file in files_in_row:
            src_path = find_file_in_sources(target_file)
            if src_path:
                try:
                    eq_folder_name = extract_eqname(src_path)
                    break
                except Exception as e:
                    print(f"⚠️ Metadata extraction failed for {target_file}: {e}")
                    continue

        if not eq_folder_name:
            eq_folder_name = f"Row_{idx + 1}_Unknown"

        eq_folder = os.path.join(destination_root, eq_folder_name)
        os.makedirs(eq_folder, exist_ok=True)
        print(f"\n📂 Created folder: {eq_folder_name}")

        # --- Copy all files from this row ---
        for target_file in files_in_row:
            src_path = find_file_in_sources(target_file)
            if src_path:
                shutil.copy2(src_path, os.path.join(eq_folder, os.path.basename(src_path)))
                print(f"   ✅ Copied: {target_file}")
            else:
                print(f"   ⚠️ File not found in any source folder: {target_file}")

    print("\n🎉 Grouping complete using multiple source folders and metadata-based naming!")
    print(f"📁 Output saved to: {destination_root}")


def EQ_H_V_Segregation(destination_root):
    earthquake_files = {}
    for eq_folder in os.listdir(destination_root):
        eq_path = os.path.join(destination_root, eq_folder)
        if not os.path.isdir(eq_path):
            continue
        earthquake_files[eq_folder] = {"Horizontal": [], "Vertical": []}


        for file in os.listdir(eq_path):
            if file.lower().endswith(".at2"):
                # Extract last 3 letters before extension
                name_part = os.path.splitext(file)[0]  # filename without extension
                last3 = name_part[-3:].upper()  # last 3 letters, uppercase
                file_path = os.path.join(eq_path, file)


                if any(x in file.upper() for x in ["DWN", "UP", "V1"]):
                    earthquake_files[eq_folder]["Vertical"].append(file_path)

                else:
                    earthquake_files[eq_folder]["Horizontal"].append(file_path)

    # for key, value in earthquake_files.items():
    #     print (key)
    #     for i,  j in value.items():
    #         print(i)
    #         print(j)
    return earthquake_files


def RotDnn_Matching(seed1, seed2, target_file):
    """
    Example 3: Direct RotDnn Component Matching (Concise)

    Modifies two horizontal components from a historic record simultaneously so that
    the resulting RotD100 response spectrum (computed from the pair)
    matches the specified RotD100 design/target spectrum.

    This is the recommended approach for matching two components.

    """

    plt.close('all')

    # --- Configuration ---
    # Setup basic logging to see output from the module
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')


    seed_file_1 = seed1  # Seed record comp1 [g]
    seed_file_2 = seed2  # Seed record comp2 [g]
    target_file = target_file  # Target spectrum (T, PSA)
    dampratio = 0.0299  # Damping ratio for spectra
    TL1 = 0.05  # Lower period limit for matching (s)
    TL2 = 6.0  # Upper period limit for matching (s)
    nit_match = 15
    nn = 100  # Percentile for RotD (100 = RotD100)
    baseline_correct = True
    p_order = -1
    output_base_name = '_RotD' + str(nn) + os.path.basename(seed1)[:-10] + '_' + os.path.basename(target_file)[:-4]   # Base name for output files

    # --- Load target spectrum and seed record ---

    s1, dt, n1, name1 = load_PEERNGA_record(seed_file_1)
    s2, _, n2, name2 = load_PEERNGA_record(seed_file_2)

    fs = 1 / dt

    target_spectrum = np.loadtxt(target_file)
    sort_idx = np.argsort(target_spectrum[:, 0])
    To = target_spectrum[sort_idx, 0]  # Target spectrum periods
    dso = target_spectrum[sort_idx, 1]  # Target spectrum PSA

    # --- Perform Direct RotDnn Spectral Matching ---
    # Call the  REQPYrotdnn function
    results = REQPYrotdnn(
        s1=s1,
        s2=s2,
        fs=fs,
        dso=dso,
        To=To,
        nn=nn,
        T1=TL1,
        T2=TL2,
        zi=dampratio,
        nit=nit_match,
        baseline=baseline_correct,
        porder=p_order)

    print("Spectral matching complete.")
    print(f"Final RMSE (pre-BC): {results.get('rmsefin', 'N/A'):.2f}%")
    print(f"Final Misfit (pre-BC): {results.get('meanefin', 'N/A'):.2f}%")

    # --- Plot Results ---
    # Call the plotting function for RotDnn results
    fig_hist, fig_spec = plot_rotdnn_results(
        results=results,
        s1_orig=s1,  # Pass original unscaled record 1
        s2_orig=s2,  # Pass original unscaled record 2
        target_spec=(To, dso),
        T1=TL1,
        T2=TL2,
        xlim_min=None,
        xlim_max=None)

    # Save and show plots
    directory = os.path.dirname(seed1)

    hist_filename = f"{output_base_name}_TimeHistories.png"
    hist_filename = os.path.join(directory, hist_filename)

    spec_filename = f"{output_base_name}_Spectra.png"
    spec_filename = os.path.join(directory, spec_filename)

    fig_hist.savefig(hist_filename, dpi=300)
    fig_spec.savefig(spec_filename, dpi=300)
    print(f"Saved plots to {hist_filename} and {spec_filename}")
    plt.show()

    # --- Save Matched Records ------------------------------------------------------------------------------------

    # --- Save Component 1 ---

    at2_filepath1 = f"{output_base_name}_Comp1_Matched.AT2"
    at2_filepath1 = os.path.join(directory, at2_filepath1)

    at2_header1 = {
        'title': f'Matched record from {seed_file_1} (Target: {target_file})',
        'station': name1.split('_comp_')[0] if '_comp_' in name1 else name1,
        'component': f"{name1.split('_comp_')[-1]}-Matched"
    }
    save_results_as_at2(results, at2_filepath1, comp_key='scc1', header_details=at2_header1)

    txt_1col_filepath1 = f"{output_base_name}_Comp1_Matched_1col.txt"
    txt_1col_filepath1 = os.path.join(directory, txt_1col_filepath1)
    header_1col_1 = (f"Matched acceleration (g), dt={results.get('dt', 0.0):.8f}s\n"
                     f"Original Seed: {name1}\n"
                     f"Target Spectrum: {target_file}\n"
                     f"Data points follow:")
    save_results_as_1col(results, txt_1col_filepath1, comp_key='scc1', header_str=header_1col_1)

    # --- Save Component 2 ---
    at2_filepath2 = f"{output_base_name}_Comp2_Matched.AT2"
    at2_filepath2 = os.path.join(directory, at2_filepath2)

    at2_header2 = {
        'title': f'Matched record from {seed_file_2} (Target: {target_file})',
        'station': name2.split('_comp_')[0] if '_comp_' in name2 else name2,
        'component': f"{name2.split('_comp_')[-1]}-Matched"
    }
    save_results_as_at2(results, at2_filepath2, comp_key='scc2', header_details=at2_header2)

    txt_1col_filepath2 = f"{output_base_name}_Comp2_Matched_1col.txt"
    txt_1col_filepath2 = os.path.join(directory, txt_1col_filepath2)

    header_1col_2 = (f"Matched acceleration (g), dt={results.get('dt', 0.0):.8f}s\n"
                     f"Original Seed: {name2}\n"
                     f"Target Spectrum: {target_file}\n"
                     f"Data points follow:")
    save_results_as_1col(results, txt_1col_filepath2, comp_key='scc2', header_str=header_1col_2)

    print(f"Saved {at2_filepath1}, {txt_1col_filepath1}, {at2_filepath2}, {txt_1col_filepath2}")

    print("\nScript finished.")

def SingleComp_Matching(seed1, target_file, showPlot = False):
    """
    Example 1: Single Component Spectral Matching (Concise)

    Matches a single component to a target spectrum using the refactored module.
    Removes logging and defensive error handling around core functions for brevity.
    """

    # Import necessary functions f

    plt.close('all')
    # --- Configuration ---
    # Setup basic logging to see output from the module
    logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

    seed_file = seed1  # Seed record [g]
    target_file = target_file  # Target spectrum (T, PSA)
    dampratio = 0.05  # Damping ratio for spectra
    TL1 = 0.05  # Lower period limit for matching (s)
    TL2 = 6.0  # Upper period limit for matching (s)
    nit_match = 15  # Number of matching iterations
    baseline_correct = True  # Perform baseline correction?
    p_order = -1  # Detrending order for baseline (-1 = none)

    output_base_name = "_SingleMatch_" + os.path.basename(seed_file)[:-4] + '_' + os.path.basename(target_file)[:-4]  # Base name for output files
    directory = os.path.dirname(seed1)

    # --- Load target spectrum and seed record ---

    s_orig, dt, npts, eqname = load_PEERNGA_record(seed_file)
    fs = 1 / dt

    target_spectrum = np.loadtxt(target_file)
    if target_spectrum.ndim != 2 or target_spectrum.shape[1] != 2:
        raise ValueError("Target file should have two columns (Period, PSA).")

    sort_idx = np.argsort(target_spectrum[:, 0])
    To = target_spectrum[sort_idx, 0]  # Target spectrum periods
    dso = target_spectrum[sort_idx, 1]  # Target spectrum PSA

    # --- Perform Spectral Matching ---
    results = REQPY_single(
        s=s_orig,
        fs=fs,
        dso=dso,
        To=To,
        T1=TL1,
        T2=TL2,
        zi=dampratio,
        nit=nit_match,
        baseline=baseline_correct,
        porder=p_order)

    print("Spectral matching complete.")
    print(f"Final RMSE (pre-BC): {results['rmsefin']:.2f}%")
    print(f"Final Misfit (pre-BC): {results['meanefin']:.2f}%")

    # --- Extract Results ---
    ccs = results['ccs']
    cvel = results['cvel']
    cdespl = results['cdespl']

    # --- Plot Results ---

    fig_hist, fig_spec = plot_single_results(
        results=results,
        s_orig=s_orig,
        target_spec=(To, dso),
        T1=TL1,
        T2=TL2,
        xlim_min=None,
        xlim_max=None)

    # Save and show plots


    hist_filename = f"{output_base_name}_TimeHistories.png"
    hist_filename = os.path.join(directory, hist_filename)

    spec_filename = f"{output_base_name}_Spectra.png"
    spec_filename = os.path.join(directory, spec_filename)

    fig_hist.savefig(hist_filename, dpi=300)
    fig_spec.savefig(spec_filename, dpi=300)

    if showPlot:
        plt.show()  # Display plots

    # --- Save Matched Record ---

    # --- Option 1: Save as .AT2 format ---
    at2_filepath = f"{output_base_name}_Matched.AT2"
    at2_filepath = os.path.join(directory, at2_filepath)

    at2_header_details = {
        'title': f'Matched record from {seed_file} (Target: {target_file})',
        'date': '01/01/2025',  # Placeholder date
        'station': eqname.split('_comp_')[0] if '_comp_' in eqname else eqname,
        'component': f"{eqname.split('_comp_')[-1]}-Matched"
    }
    save_results_as_at2(results, at2_filepath, comp_key='ccs', header_details=at2_header_details)

    # --- Option 2: Save as 2-column (Time, Accel) .txt file ---
    txt_2col_filepath = f"{output_base_name}_Matched_2col.txt"
    txt_2col_filepath = os.path.join(directory, txt_2col_filepath)

    header_2col = (f"Matched acceleration (g) vs. Time (s)\n"
                   f"Original Seed: {eqname}\n"
                   f"Target Spectrum: {target_file}\n"
                   f"Time (s), Acceleration (g)")
    save_results_as_2col(results, txt_2col_filepath, comp_key='ccs', header_str=header_2col)

    # --- Option 3: Save as 1-column (Accel) .txt file ---
    txt_1col_filepath = f"{output_base_name}_Matched_1col.txt"
    txt_1col_filepath = os.path.join(directory, txt_1col_filepath)

    header_1col = (f"Matched acceleration (g), dt={results.get('dt', 0.0):.8f}s\n"
                   f"Original Seed: {eqname}\n"
                   f"Target Spectrum: {target_file}\n"
                   f"Data points follow:")
    save_results_as_1col(results, txt_1col_filepath, comp_key='ccs', header_str=header_1col)

    print("\nScript finished.")
    return at2_filepath

def ResponseSpectrum(accSeries, dt, showPlot = False):
    # !/usr/bin/env python3
    # -*- coding: utf-8 -*-
    """
    Created on Mon Jul 27 09:55:02 2020

    @author: Loic Viens
    Example to compute the response spectra from acceleration waveforms. The data are from the 2011 Mw 9.0 Tohoku-Oki earthquake recorded at the TKYH12 KiK-net station.

    - Resp_type: Response type, choose between:
                 - 'SA'  : Acceleration Spectra
                 - 'PSA' : Pseudo-acceleration Spectra
                 - 'SV'  : Velocity Spectra
                 - 'PSV' : Pseudo-velocity Spectra
                 - 'SD'  : Displacement Spectra
    """

    # %% Parameters of the response spectra
    Resp_type = 'SA'  # See above for the different options
    T = np.concatenate((np.arange(0.0001, 0.01, 0.001), np.arange(0.01, 0.1, 0.005), np.arange(0.1, 0.5, 0.01), np.arange(0.5, 1, 0.02),
                        np.arange(1, 5, 0.1), np.arange(5, 15.5, .5)))  # Time vector for the spectral response
    freq = 1 / T  # Frequenxy vector
    xi = .05  # Damping factor
    delta = 1 / dt
    Sfin = RS_function(accSeries, delta, T, xi, Resp_type=Resp_type)

    #Timeseries data
    t = np.linspace(0, len(accSeries) * dt, len(accSeries))  # Time vector
    accSeries = accSeries
    timeseries = {"accSeries": accSeries, "t":t}


    #Response Spectrum Data
    Rsdata = {"Resp_type":Resp_type,"Sfin":Sfin, "T": T, "freq": freq, "xi": xi}

    return timeseries, Rsdata


def fix_RSperiodspacing(file):
    import numpy as np
    from scipy.interpolate import interp1d

    # Read data from the file
    data = np.loadtxt(file)  # Replace with your file path

    # Separate columns
    x = data[:, 0]
    y = data[:, 1]

    # Create interpolation function
    f = interp1d(x, y, kind='cubic')  # linear interpolation; you can use 'cubic' too

    # Create new x values with interval 0.001
    x_new = np.arange(x[0], x[-3] + 0.001, 0.001)

    # Interpolate to get new y values
    y_new = f(x_new)

    # Save interpolated data to a new file
    np.savetxt("A_interpolated.txt", np.column_stack((x_new, y_new)), fmt='%.6f', delimiter='\t')

    print("Interpolation done! Data saved to 'A_interpolated.txt'.")

#===============================================================RS period adjustments
# fix_RSperiodspacing(r"C:\Users\Acer\Documents\Civil_Multipurpose Software\IS1893_Soft.txt")

#===============================================================EQs Grouping
folderpath = r'E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\magnitude of 6 to 7.5\300to 800 Vs rescords'
# earthquake_groups = EarthquakeGrouping(folderpath)

#===============================================================EQs Extraction from dir based on Excel selected rows
source_folders = [
    r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\magnitude of 6 to 7.5\300to 800 Vs rescords",
    r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\magnitude of 6 to 7.5\200to 300",
    r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\magnitude of 6 to 7.5\Selected"
]
excel_path = r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\magnitude of 6 to 7.5\Earthquakes Data.xlsx"
sheet_name = "Selected EQS"
destination_root = r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\Grouped_By_Row_Metadata"
# GrouptoFolder_FromSelectedinExcel(source_folders, excel_path, sheet_name, destination_root)

#===============================================================EQs Segregation to horizontal and vertical components files
destination_root = r"E:\Machine Learning Research\Numerical Analysis\Earthquakes Materials\Grouped_By_Row_Metadata"
earthquake_files = EQ_H_V_Segregation(destination_root)

#===============================================================Matching and handling
def DataProcessing(earthquake_files):
    def constIntTS(timeseries, intervals):
        t = np.asarray(timeseries["t"], dtype=float)
        accSeries = timeseries["accSeries"]
        accSeries = np.asarray(accSeries, dtype=float)

        # Create uniform time array from min to max
        t_const = np.arange(t[0], t[-1] + intervals, intervals)

        # Interpolate acceleration at uniform time points
        accSeriesScaledConstInt = np.interp(t_const, t, accSeries)
        timeseries = {"accSeries": accSeriesScaledConstInt, "t": t_const}
        return timeseries
    GlobalDict = {}

    Plot = False

    SaveExcel = True

    Scaling = True
    refPeriod = 0.55

    tsConstInterval = True
    intervals = 0.01

    Unit = "m/s²"

    ii = 1
    for Earthquake, InDict in earthquake_files.items():
        Horizontals = InDict["Horizontal"]
        Verticals = InDict["Vertical"]

        seed1 = Horizontals[0]
        seed2 = Horizontals[1]
        EQNameOnly = extract_eqname(seed1, EqNameOnly = True)

        acc1, dt1, npts1, eqname1 = load_PEERNGA_record(seed1)
        acc2, dt2, npts2, eqname2 = load_PEERNGA_record(seed2)

        # Maximum values
        max_acc1 = np.max(acc1)
        max_acc2 = np.max(acc2)

        # Determine overall maximum and its source
        if max_acc1 >= max_acc2:
            overall_max = max_acc1
            source_seed = seed1
            line3 = ["Selected(max g):", max_acc1, dt1, npts1, eqname1]

        else:
            overall_max = max_acc2
            source_seed = seed2
            line3 = ["Selected(max g):", max_acc1, dt1, npts1, eqname1]
        acc, dt, npts, eqname = load_PEERNGA_record(source_seed)


        directory = os.path.dirname(seed1)
        line1 = [ max_acc1, dt1, npts1, eqname1]
        line2 = [ max_acc2, dt2, npts2, eqname2]

        # Write lines to the file
        output_file = os.path.join(directory, "_details.txt")
        with open(output_file, "w") as f:
            f.write("\t".join(map(str, line1)) + "\n")
            f.write("\t".join(map(str, line2)) + "\n")
            f.write("\t".join(map(str, line3)) + "\n")

        target_file = r"C:\Users\Acer\Documents\Civil_Multipurpose Software\IS1893_Soft.txt"

    #==============================================================Response Spectrum Creation
        timeseries, Rsdata = ResponseSpectrum(acc, dt, showPlot=False)
        if EQNameOnly not in GlobalDict:
            GlobalDict[EQNameOnly] = {}

        GlobalDict[EQNameOnly]["timeseries"] = timeseries
        GlobalDict[EQNameOnly]["Rsdata"] = Rsdata
        # Rsdata = {"Resp_type":Resp_type,"Sfin":Sfin, "T": T, "freq": freq}
        # timeseries = {"accSeries": accSeries, "t":t}

        ##_____________________Scaling of RS data
        if Scaling:
            RS_acclns = GlobalDict[EQNameOnly]["Rsdata"]["Sfin"]
            RS_periods = GlobalDict[EQNameOnly]["Rsdata"]["T"]
            RS_at_ref = np.interp(refPeriod, RS_periods, RS_acclns)
            ScaleFactor = 1/RS_at_ref

            accSeriesScaled = GlobalDict[EQNameOnly]["timeseries"]["accSeries"] * ScaleFactor
            timeseries, Rsdata = ResponseSpectrum(accSeriesScaled, dt1, showPlot=False)

            #Constant interval for all time periods of EQs
            if tsConstInterval:
                timeseries = constIntTS(timeseries, intervals)
                dt = intervals
            #Unit
            if Unit == "m/s²":
                timeseries["accSeries"] = timeseries["accSeries"] * 9.81
                accSeriesScaled = timeseries["accSeries"]



            if EQNameOnly not in GlobalDict:
                GlobalDict[EQNameOnly] = {}
            GlobalDict[EQNameOnly]["timeseries"] = timeseries
            GlobalDict[EQNameOnly]["Rsdata"] = Rsdata

            output_base_name = "_Scaled_" + os.path.basename(source_seed)[:-4] + '_' + "RefPeriod" + '_' +  str(refPeriod)  +".AT2"      # Base name for output files
            directory = os.path.join(os.path.dirname(source_seed), "Scaled")
            if not os.path.exists(directory):
                os.mkdir(directory)



            at2_filepath = os.path.join(directory, output_base_name)
            at2_header_details = {    'title': f'Matched record from {source_seed} (Scaled(RefPeriod): {refPeriod})',
                                      'date': '01/01/2025', # Placeholder date
                                      'station': eqname.split('_comp_')[0] if '_comp_' in eqname else eqname,
                                      'component': f"{eqname.split('_comp_')[-1]}-Matched"}
            results = {"comp_key":accSeriesScaled, "dt":dt}
            save_results_as_at2(results, at2_filepath, comp_key='comp_key', header_details=at2_header_details, acc_format=Unit)
            print(timeseries["accSeries"])
            print(accSeriesScaled)

        # ii+=1
        # if ii == 2:
        #     break

        if tsConstInterval:
            timeseries = constIntTS(timeseries, intervals)
            GlobalDict[EQNameOnly]["timeseries"] = timeseries
        if Unit == "m/s²" and not Scaling:
            timeseries["accSeries"] = timeseries["accSeries"] * 9.81
            GlobalDict[EQNameOnly]["timeseries"] = timeseries

            output_base_name = "_UnScaled_" + os.path.basename(source_seed)[:-4]  +".AT2"      # Base name for output files
            directory = os.path.join(os.path.dirname(source_seed), "UnScaled")
            if not os.path.exists(directory):
                os.mkdir(directory)
            at2_filepath = os.path.join(directory, output_base_name)
            at2_header_details = {    'title': f'Matched record from {source_seed}',
                                      'date': '01/01/2025', # Placeholder date
                                      'station': eqname.split('_comp_')[0] if '_comp_' in eqname else eqname,
                                      'component': f"{eqname.split('_comp_')[-1]}-Matched"}
            results = {"comp_key":timeseries["accSeries"], "dt":dt}
            save_results_as_at2(results, at2_filepath, comp_key='comp_key', header_details=at2_header_details, acc_format=Unit)
            print(timeseries["accSeries"])
            print(accSeriesScaled)


    #==============================================================Single Wave matching to the RS
        # accfile = SingleComp_Matching(source_seed, target_file)


    #==============================================================RotDnn Matching to the RS
        # RotDnn_Matching(seed1, seed2, target_file)


    #==============================================================Plotting for the data
    EarthquakeNames = []
    Resp_types = []
    xis = []

    TS_X = []
    TS_Y = []

    RS_X = []
    RS_Y = []

    Fr_X = []
    for eq, data in GlobalDict.items():
        ts_y = data["timeseries"]["accSeries"]
        ts_x = data["timeseries"]["t"]
        rs_y = data["Rsdata"]['Sfin']
        rs_x = data["Rsdata"]['T']

        TS_X.append(ts_x)
        TS_Y.append(ts_y)
        RS_X.append(rs_x)
        RS_Y.append(rs_y)

        EarthquakeNames.append(eq)
        Resp_types.append(data["Rsdata"]['Resp_type'])
        xis.append(data["Rsdata"]['xi'])
        Fr_X.append(data["Rsdata"]['freq'])

    if Plot:
        # Plot time series
        fig = plt.figure(figsize=(6, 10))
        plt.subplot(3, 1, 1)
        for index, tseries in enumerate(TS_X):
            plt.plot(TS_X[index], TS_Y[index], label=EarthquakeNames[index])
        plt.grid()
        plt.title('Earthquake Timeseries')
        plt.legend(loc=1)
        plt.ylabel(f'Acceleration ({Unit})')
        plt.xlabel('Time (s)')
        plt.xlim(TS_X[0][0], TS_X[0][-1])

        #
        # Plot response spectra with period axis
        plt.subplot(3, 1, 2)
        for index, tseries in enumerate(RS_X):
            plt.semilogy(RS_X[index], RS_Y[index], linewidth=2, label=EarthquakeNames[index])
        plt.grid()
        plt.title('Damping: ' + str(xis[0] * 100) + ' %')
        plt.legend(loc=1)
        if Resp_types[index] == 'SA':
            plt.ylabel('Acceleration response (g)')
        elif Resp_types[index]  == 'SV':
            plt.ylabel('Velocity response (cm/s)')
        else:
            plt.ylabel(Resp_types[index] )
        plt.xlabel('Period (s)')
        plt.xlim(RS_X[0][0], RS_X[0][-1])

        # Plot response spectra with frequency axis
        plt.subplot(3, 1, 3)
        for index, tseries in enumerate(RS_X):
            plt.semilogy(Fr_X[index], RS_Y[index], linewidth=2, label=EarthquakeNames[index])
        plt.grid()
        plt.xlim(Fr_X[0][0], Fr_X[0][-1])
        plt.legend(loc=4)
        if Resp_types[index] == 'SA':
            plt.ylabel('Acceleration response (g)')
        elif Resp_types[index]  == 'SV':
            plt.ylabel('Velocity response (cm/s)')
        else:
            plt.ylabel(Resp_types[index] )
        plt.xlabel('Frequency (Hz)')
        plt.tight_layout()
        plt.show()

    if SaveExcel:
        # You already have these lists:
        # EarthquakeNames, Resp_types, xis, TS_X, TS_Y, RS_X, RS_Y, Fr_X

        save_path = r"C:\Users\Acer\Documents\Civil_Multipurpose Software\Earthquake_Data_Merged.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "All_Earthquakes"

        col_cursor = 1  # starting column for first earthquake
        time_series_positions = []
        rs_positions = []
        fr_positions = []

        for i, eq_name in enumerate(EarthquakeNames):
            row_cursor = 1  # each block starts from top row

            # ===== Title =====
            title = f"Earthquake: {eq_name} | Response Type: {Resp_types[i]} | Damping (xi): {xis[i]}"
            ws.merge_cells(start_row=row_cursor, start_column=col_cursor,
                           end_row=row_cursor, end_column=col_cursor + 5)
            cell = ws.cell(row=row_cursor, column=col_cursor, value=title)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            row_cursor += 2

            # ===== Time Series =====
            datastartrow = row_cursor
            ws.cell(row=row_cursor, column=col_cursor, value="Time Series Data")
            ws.cell(row=row_cursor, column=col_cursor).font = Font(bold=True, underline="single")
            row_cursor += 1

            ts_df = pd.DataFrame({
                "Time (s)": TS_X[i],
                f"Acceleration ({Unit})": TS_Y[i]
            })

            for r_idx, row in enumerate(dataframe_to_rows(ts_df, index=False, header=True), start=row_cursor):
                for c_idx, val in enumerate(row, start=col_cursor):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            ts_start = row_cursor + 1
            ts_end = row_cursor + len(ts_df)
            time_series_positions.append((col_cursor, ts_start, ts_end, eq_name))
            row_cursor += len(ts_df) + 2

            # ===== Response Spectrum =====
            col_cursor  =col_cursor +2
            row_cursor = datastartrow
            ws.cell(row=row_cursor, column=col_cursor, value="Response Spectrum Data")
            ws.cell(row=row_cursor, column=col_cursor).font = Font(bold=True, underline="single")
            row_cursor += 1

            rs_df = pd.DataFrame({
                "Period (s)": RS_X[i],
                "Spectral Acc (g)": RS_Y[i]
            })

            for r_idx, row in enumerate(dataframe_to_rows(rs_df, index=False, header=True), start=row_cursor):
                for c_idx, val in enumerate(row, start=col_cursor):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            rs_start = row_cursor + 1
            rs_end = row_cursor + len(rs_df)
            rs_positions.append((col_cursor, rs_start, rs_end, eq_name))
            row_cursor += len(rs_df) + 2

            # ===== Frequency Response =====
            col_cursor  =col_cursor +2
            row_cursor = datastartrow
            ws.cell(row=row_cursor, column=col_cursor, value="Frequency Response Data")
            ws.cell(row=row_cursor, column=col_cursor).font = Font(bold=True, underline="single")
            row_cursor += 1

            fr_df = pd.DataFrame({
                "Frequency (Hz)": Fr_X[i],
                "Spectral Acc (g)": RS_Y[i]
            })

            for r_idx, row in enumerate(dataframe_to_rows(fr_df, index=False, header=True), start=row_cursor):
                for c_idx, val in enumerate(row, start=col_cursor):
                    ws.cell(row=r_idx, column=c_idx, value=val)
            fr_start = row_cursor + 1
            fr_end = row_cursor + len(fr_df)
            fr_positions.append((col_cursor, fr_start, fr_end, eq_name))

            # move start column for next earthquake (leave one empty column)
            col_cursor = ws.max_column + 2

        def add_chart(ws, title, x_label, y_label, positions, x_offset=0, y_offset=1, anchor="A80", log_scale=False):
            chart = ScatterChart()
            chart.title = title
            chart.x_axis.title = x_label
            chart.y_axis.title = y_label

            if log_scale:
                chart.y_axis.scaling.logBase = 10


            for col, start, end, name in positions:
                x_ref = Reference(ws, min_col=col + x_offset, min_row=start, max_row=end)
                y_ref = Reference(ws, min_col=col + y_offset, min_row=start, max_row=end)
                series = openpyxl.chart.Series(y_ref, x_ref, title=name)
                chart.series.append(series)


            pp = ParagraphProperties(defRPr = CharacterProperties(latin=DrawFont(typeface='Times New Roman'), sz=900, b=False))
            rtp = RichText(p=[Paragraph(pPr=pp,
                                        endParaRPr=CharacterProperties(latin=DrawFont(typeface='Times New Roman'), sz=900,
                                                                       b=False))])
            chart.x_axis.txPr = rtp
            chart.y_axis.txPr = rtp
            pp = ParagraphProperties(
                defRPr=CharacterProperties(latin=DrawFont(typeface='Times New Roman'), sz=1000, b=True))
            rtp = RichText(p=[Paragraph(pPr=pp,
                                        endParaRPr=CharacterProperties(latin=DrawFont(typeface='Times New Roman'), sz=1000,
                                                                       b=True))])
            chart.legend.txPr = rtp
            chart.title.tx.rich.p[0].pPr = pp
            chart.x_axis.title.tx.rich.p[0].pPr = pp
            chart.y_axis.title.tx.rich.p[0].pPr = pp
            # Adjust graph size
            chart.width = 24  # Elselvier page halfwidth = 9cm        Writing area Only
            chart.height = 12  # Elselvier page full height = 24 cm        Writing area Only

            ws.add_chart(chart, anchor)
        # ===== Add 3 charts =====
        add_chart(ws, "Time Series Comparison", "Time (s)", f"Acceleration ({Unit})", time_series_positions, anchor="B10")
        add_chart(ws, f"Response Spectrum Comparison (Damping {xis[0] * 100}%)", "Period (s)", "Spectral Acc (g)",
                  rs_positions, anchor="O10", log_scale=True)
        add_chart(ws, f"Frequency Response Comparison (Damping {xis[0] * 100}%)", "Frequency (Hz)",
                  "Spectral Acc (g)", fr_positions, anchor="AE10", log_scale=True)


        wb.save(save_path)
        print(f"✅ All earthquake data written side-by-side to:\n{save_path}")


DataProcessing(earthquake_files)


